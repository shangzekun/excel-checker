import io
import math
import logging
from pathlib import Path
from collections import defaultdict
from typing import List, Dict, Any, Callable, Optional

from openpyxl import load_workbook

from .schemas import Issue

logger = logging.getLogger(__name__)

TARGET_SHEETS = ["RSW", "SPR", "FDS"]
MATERIAL_DEPOT_PATH = Path("data/material_depot.xlsx")
MATERIAL_DEPOT_SHEET = "material_depot"
RULE_NAME_MAP = {
    "input": "数据成熟度检查",
    "P1": "BOM重复零件冲突检查",
    "rule1": "连接点号规范性检查",
    "rule2": "材料牌号规范性检查",
    "rule3": "连接点号重复性检查",
    "rule4": "连接点号位置重复性检查",
    "rule5": "钣金零件：同一零件版本号一致性检查",
    "rule6": "钣金零件：同一零件材料牌号一致性检查",
    "rule7": "钣金零件：同一零件材料厚度一致性检查",
    "rule8": "Fixing-Part版本号与BOM中版本号一致性检查",
    "rule9": "零件版本号：点表与BOM一致性检查",
    "rule10": "零件材料牌号：点表与BOM一致性检查",
    "rule11": "零件厚度：点表与BOM一致性检查",
    "rule12": "工艺点属性的检查",
    "rule8-11": "BOM一致性检查",
}
ROOT_RULE_ID = "data_maturity_check"

def _display_rule_name(rule_code: str) -> str:
    return RULE_NAME_MAP.get(rule_code, rule_code)

def get_rules() -> List[Dict[str, Any]]:
    return [
        {
            "id": "data_maturity_check",
            "name": "数据成熟度检查",
            "level": "warning",
            "description": "",
            "enabled": True,
            "details": [
                "连接点号规范性检查：连接点号为“xxx123456_7890”，其中，“xxx”为对应工艺的三位字母，123456代表Fixing Part的后六位数字，最后五位为“下划线”+四位数字。标准点号规范如：“RSW349075_0104”。",
                "材料牌号规范性检查：建立了标准材料牌号规范库，对点表中的材料牌号进行一一比对，判断点表中材料牌号是否规范。",
                "连接点号重复性检查：依次遍历点表中的连接点号，判断连接点号是否唯一。",
                "连接点号位置重复性检查：按照空间位置10mm距离进行检查，两连接点的空间位置相距≤10mm报错。",
                "钣金零件：同一零件版本号一致性检查：依次遍历点表中的零件及其版本号，判断同一零件在该点表中的版本号是否唯一。",
                "钣金零件：同一零件材料牌号一致性检查：依次遍历点表中的零件及其材料牌号，判断同一零件在该点表中的材料牌号是否唯一。",
                "钣金零件：同一零件材料厚度一致性检查：依次遍历点表中的零件及其厚度，判断同一零件在该点表中的厚度是否唯一。由于铸件的厚度不唯一，该检查项并没有对铸件的厚度进行检查。",
                "Fixing-Part版本号与BOM中版本号一致性检查：将点表中Fixing-Part的版本号与选中的BOM清单中的版本号进行比对，判断两者是否相同。",
                "零件版本号：点表与BOM一致性检查：将点表中每个零件的版本号与选中的BOM清单中的该零件版本号进行比对，判断两者是否相同。",
                "零件材料牌号：点表与BOM一致性检查：将点表中每个零件的材料牌号与选中的BOM清单中的该零件材料牌号进行比对，判断两者是否相同。",
                "零件厚度：点表与BOM一致性检查：将点表中每个零件的厚度与选中的BOM清单中的该零件厚度进行比对，判断两者是否相同。由于铸件的厚度不唯一，该检查项并没有对铸件的厚度进行检查。",
                "工艺点属性的检查：如果该连接点属于单边点焊或四层搭接，判断该连接点是否被标注为工艺点。如果该连接点被标注为工艺点，判断该连接点是否属于单边点焊或四层搭接，如若不是给出提示，需人工确认该点是否为工艺点。",
            ],
        },
        {
            "id": "required_fields_demo",
            "name": "必填字段检查（Demo）",
            "level": "error",
            "description": "模拟检查关键字段是否为空，返回若干 error 示例。",
            "enabled": False,
            "details": ["当前保留为示例规则，默认关闭。"],
        },
        {
            "id": "date_format_demo",
            "name": "日期格式检查（Demo）",
            "level": "warning",
            "description": "模拟检查日期格式是否标准，返回若干 warning 示例。",
            "enabled": False,
            "details": ["当前保留为示例规则，默认关闭。"],
        },
        {
            "id": "enum_value_demo",
            "name": "枚举值检查（Demo）",
            "level": "info",
            "description": "模拟检查枚举值是否在允许范围内，返回若干 info 示例。",
            "enabled": False,
            "details": ["当前保留为示例规则，默认关闭。"],
        },
        {
            "id": "cross_sheet_demo",
            "name": "跨表关联检查（Demo）",
            "level": "error",
            "description": "模拟主子表关联关系检查，例如主键缺失、引用不存在等。",
            "enabled": False,
            "details": ["当前保留为示例规则，默认关闭。"],
        },
        {
            "id": "range_check_demo",
            "name": "数值范围检查（Demo）",
            "level": "warning",
            "description": "模拟检查金额、数量、比例等是否落在合理区间。",
            "enabled": False,
            "details": ["当前保留为示例规则，默认关闭。"],
        },
    ]


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: Any) -> str:
    return _normalize_text(value).lower()


def _safe_float(value: Any) -> Optional[float]:
    text = _normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _norm_num_text(value: Any) -> str:
    f = _safe_float(value)
    if f is None:
        return _normalize_text(value)
    if float(f).is_integer():
        return str(int(f))
    return f"{f:.6f}".rstrip("0").rstrip(".")


def _split_bom_gauge(value: Any) -> List[str]:
    text = _normalize_text(value)
    if not text:
        return []
    parts = [x.strip() for x in text.split("/") if x.strip()]
    return [_norm_num_text(x) for x in parts]


def _format_values(values: List[str] | set[str]) -> str:
    return " / ".join(sorted({_normalize_text(v) for v in values if _normalize_text(v)}))


def _compose_bom_version(part_rev: str, minor_rev: str) -> str:
    part_rev = _normalize_text(part_rev)
    minor_rev = _normalize_text(minor_rev)
    if part_rev and minor_rev:
        return f"{part_rev}.{minor_rev}"
    if part_rev:
        return part_rev
    if minor_rev:
        return minor_rev
    return ""


def _load_workbook_from_bytes(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), data_only=True)


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        key = _normalize_key(cell.value)
        if key and key not in mapping:
            mapping[key] = col_idx
    return mapping


def _cell_value(ws, row_idx: int, header_map: Dict[str, int], header_name: str) -> Any:
    col_idx = header_map.get(_normalize_key(header_name))
    if not col_idx:
        return None
    return ws.cell(row=row_idx, column=col_idx).value


def _issue(
    level: str,
    message: str,
    sheet: Optional[str],
    row: Optional[int],
    column: Optional[str],
    rule: str,
    group_key: Optional[str] = None,
    group_title: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_value: Optional[str] = None,
    details: Optional[Dict[str, Any]] = None,
) -> Issue:
    return Issue(
        level=level,
        message=message,
        sheet=sheet,
        row=row,
        column=column,
        rule=rule,
        group_key=group_key,
        group_title=group_title,
        entity_type=entity_type,
        entity_value=entity_value,
        details=details or {},
    )


def _collect_required_headers(ws, header_row: int, required_headers: List[str], rule: str) -> List[Issue]:
    issues: List[Issue] = []
    header_map = _build_header_map(ws, header_row)
    for name in required_headers:
        if _normalize_key(name) not in header_map:
            issues.append(
                _issue(
                    level="error",
                    message=f"缺少必要列：{name}",
                    sheet=ws.title,
                    row=header_row,
                    column=None,
                    rule=rule,
                    group_key=f"{rule}:missing_header:{ws.title}:{name}",
                    group_title=f"缺少必要列：{name}",
                    entity_type="header",
                    entity_value=name,
                )
            )
    return issues


def _iter_sheet_rows(ws, header_row: int, used_columns: set[int] | None = None) -> List[int]:
    rows: List[int] = []
    if used_columns:
        min_col = min(used_columns)
        max_col = max(used_columns)
    else:
        min_col = 1
        max_col = ws.max_column

    for row_idx, row_values in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if used_columns:
            relevant = [
                row_values[col - min_col]
                for col in used_columns
                if 0 <= (col - min_col) < len(row_values)
            ]
        else:
            relevant = row_values

        if any(_normalize_text(v) for v in relevant):
            rows.append(row_idx)
    return rows


def _load_material_depot() -> tuple[set[str], List[Issue]]:
    issues: List[Issue] = []

    if not MATERIAL_DEPOT_PATH.exists():
        issues.append(
            _issue(
                level="warning",
                message=f"未找到材料库文件：{MATERIAL_DEPOT_PATH.as_posix()}，rule2 已跳过。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("rule2"),
                group_key="rule2:material_depot_missing",
                group_title="材料库文件缺失，rule2 已跳过",
                entity_type="material_depot",
                entity_value=MATERIAL_DEPOT_PATH.as_posix(),
            )
        )
        return set(), issues

    try:
        wb = load_workbook(MATERIAL_DEPOT_PATH, data_only=True)
    except Exception as exc:
        issues.append(
            _issue(
                level="warning",
                message=f"材料库文件读取失败：{exc}，rule2 已跳过。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("rule2"),
                group_key="rule2:material_depot_read_failed",
                group_title="材料库文件读取失败，rule2 已跳过",
                entity_type="material_depot",
                entity_value=MATERIAL_DEPOT_PATH.as_posix(),
                details={"error": str(exc)},
            )
        )
        return set(), issues

    if MATERIAL_DEPOT_SHEET not in wb.sheetnames:
        issues.append(
            _issue(
                level="warning",
                message=f"材料库缺少 sheet：{MATERIAL_DEPOT_SHEET}，rule2 已跳过。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("rule2"),
                group_key="rule2:material_depot_sheet_missing",
                group_title="材料库 sheet 缺失，rule2 已跳过",
                entity_type="material_depot_sheet",
                entity_value=MATERIAL_DEPOT_SHEET,
            )
        )
        return set(), issues

    ws = wb[MATERIAL_DEPOT_SHEET]
    header_map = _build_header_map(ws, 1)

    material_col = None
    for candidate in ["Material", "Applied Material", "Material Name"]:
        material_col = header_map.get(_normalize_key(candidate))
        if material_col:
            break

    if not material_col:
        issues.append(
            _issue(
                level="warning",
                message="材料库缺少列：Material，rule2 已跳过。",
                sheet=MATERIAL_DEPOT_SHEET,
                row=1,
                column=None,
                rule=_display_rule_name("rule2"),
                group_key="rule2:material_column_missing",
                group_title="材料库列缺失，rule2 已跳过",
                entity_type="material_depot_header",
                entity_value="Material",
            )
        )
        return set(), issues

    materials: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=material_col).value
        text = _normalize_text(value)
        if text:
            materials.add(text)

    return materials, issues


def _load_ebom(ebom_bytes: bytes) -> tuple[dict[str, dict[str, Any]], set[str], List[Issue]]:
    issues: List[Issue] = []
    bom_lookup: dict[str, dict[str, Any]] = {}
    conflict_parts: set[str] = set()

    try:
        wb = _load_workbook_from_bytes(ebom_bytes)
    except Exception as exc:
        issues.append(
            _issue(
                level="error",
                message=f"EBOM 文件读取失败：{exc}",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("input"),
                group_key="input:ebom_read_failed",
                group_title="EBOM 文件读取失败",
                entity_type="ebom",
                details={"error": str(exc)},
            )
        )
        return bom_lookup, conflict_parts, issues

    if "EBOM" not in wb.sheetnames:
        issues.append(
            _issue(
                level="error",
                message="上传的 EBOM 文件中缺少 EBOM 页。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("input"),
                group_key="input:ebom_sheet_missing",
                group_title="上传的 EBOM 文件缺少 EBOM 页",
                entity_type="ebom_sheet",
                entity_value="EBOM",
            )
        )
        return bom_lookup, conflict_parts, issues

    ws = wb["EBOM"]
    header_row = 5
    required_headers = [
        "Part No.",
        "Part Rev",
        "Minor Rev",
        "Applied Material",
        "Gauge",
    ]

    missing = _collect_required_headers(ws, header_row, required_headers, "input")
    if missing:
        return bom_lookup, conflict_parts, missing

    header_map = _build_header_map(ws, header_row)

    raw_rows_by_part: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row_idx in _iter_sheet_rows(ws, header_row):
        part_no = _normalize_text(_cell_value(ws, row_idx, header_map, "Part No."))
        if not part_no:
            continue

        record = {
            "row": row_idx,
            "part_no": part_no,
            "part_rev": _normalize_text(_cell_value(ws, row_idx, header_map, "Part Rev")),
            "minor_rev": _normalize_text(_cell_value(ws, row_idx, header_map, "Minor Rev")),
            "applied_material": _normalize_text(_cell_value(ws, row_idx, header_map, "Applied Material")),
            "gauge_raw": _normalize_text(_cell_value(ws, row_idx, header_map, "Gauge")),
        }
        raw_rows_by_part[part_no].append(record)

    for part_no, rows in raw_rows_by_part.items():
        signatures = {
            (
                r["part_rev"],
                r["minor_rev"],
                r["applied_material"],
                r["gauge_raw"],
            )
            for r in rows
        }

        if len(signatures) > 1:
            conflict_parts.add(part_no)
            continue

        chosen = rows[0]
        bom_lookup[part_no] = {
            "part_no": part_no,
            "part_rev": chosen["part_rev"],
            "minor_rev": chosen["minor_rev"],
            "version": _compose_bom_version(chosen["part_rev"], chosen["minor_rev"]),
            "applied_material": chosen["applied_material"],
            "gauge_raw": chosen["gauge_raw"],
            "gauge_options": _split_bom_gauge(chosen["gauge_raw"]),
            "row": chosen["row"],
        }

    return bom_lookup, conflict_parts, issues


def _read_target_sheet_records(ws) -> tuple[List[dict[str, Any]], List[Issue]]:
    issues: List[Issue] = []
    header_row = 1
    required_headers = [
        "Connect ID",
        "Fixing part ID",
        "Fixing part Rev",
        "PART 1",
        "PART 2",
        "PART 3",
        "PART 4",
        "PART 1 Rev",
        "PART 2 Rev",
        "PART 3 Rev",
        "PART 4 Rev",
        "PART 1 Material",
        "PART 2 Material",
        "PART 3 Material",
        "PART 4 Material",
        "PART 1 Gauge",
        "PART 2 Gauge",
        "PART 3 Gauge",
        "PART 4 Gauge",
        "X",
        "Y",
        "Z",
        "Process Joint",
        "Extra Info",
    ]

    missing = _collect_required_headers(ws, header_row, required_headers, "input")
    if missing:
        return [], missing

    header_map = _build_header_map(ws, header_row)
    needed_columns = {
        header_map[_normalize_key(name)]
        for name in required_headers
        if _normalize_key(name) in header_map
    }
    rows = _iter_sheet_rows(ws, header_row, needed_columns)

    column_by_field = {
        "connect_id": header_map[_normalize_key("Connect ID")],
        "fixing_part_id": header_map[_normalize_key("Fixing part ID")],
        "fixing_part_rev": header_map[_normalize_key("Fixing part Rev")],
        "x": header_map[_normalize_key("X")],
        "y": header_map[_normalize_key("Y")],
        "z": header_map[_normalize_key("Z")],
        "process_joint": header_map[_normalize_key("Process Joint")],
        "extra_info": header_map[_normalize_key("Extra Info")],
    }

    part_columns: dict[int, dict[str, int]] = {}
    for i in range(1, 5):
        part_columns[i] = {
            "part_no": header_map[_normalize_key(f"PART {i}")],
            "rev": header_map[_normalize_key(f"PART {i} Rev")],
            "material": header_map[_normalize_key(f"PART {i} Material")],
            "gauge_raw": header_map[_normalize_key(f"PART {i} Gauge")],
        }

    records: List[dict[str, Any]] = []
    for row_idx in rows:
        row_map = {
            col_idx: ws.cell(row=row_idx, column=col_idx).value for col_idx in needed_columns
        }

        part_slots = []
        for i in range(1, 5):
            gauge_raw = _normalize_text(row_map.get(part_columns[i]["gauge_raw"]))
            part_slots.append(
                {
                    "slot": i,
                    "part_no": _normalize_text(row_map.get(part_columns[i]["part_no"])),
                    "rev": _normalize_text(row_map.get(part_columns[i]["rev"])),
                    "material": _normalize_text(row_map.get(part_columns[i]["material"])),
                    "gauge_raw": gauge_raw,
                    "gauge_num": _safe_float(gauge_raw),
                }
            )

        records.append(
            {
                "sheet": ws.title,
                "row": row_idx,
                "connect_id": _normalize_text(row_map.get(column_by_field["connect_id"])),
                "fixing_part_id": _normalize_text(row_map.get(column_by_field["fixing_part_id"])),
                "fixing_part_rev": _normalize_text(row_map.get(column_by_field["fixing_part_rev"])),
                "x": _safe_float(row_map.get(column_by_field["x"])),
                "y": _safe_float(row_map.get(column_by_field["y"])),
                "z": _safe_float(row_map.get(column_by_field["z"])),
                "process_joint": _normalize_text(row_map.get(column_by_field["process_joint"])),
                "extra_info": _normalize_text(row_map.get(column_by_field["extra_info"])),
                "parts": part_slots,
            }
        )

    return records, issues


def _collect_all_records(main_wb) -> tuple[List[dict[str, Any]], List[Issue], List[str]]:
    issues: List[Issue] = []
    all_records: List[dict[str, Any]] = []
    existing_sheets: List[str] = []

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in main_wb.sheetnames:
            continue
        existing_sheets.append(sheet_name)
        ws = main_wb[sheet_name]
        records, errs = _read_target_sheet_records(ws)
        issues.extend(errs)
        all_records.extend(records)

    return all_records, issues, existing_sheets


def _referenced_parts(records: List[dict[str, Any]]) -> set[str]:
    refs: set[str] = set()
    for rec in records:
        if rec["fixing_part_id"]:
            refs.add(rec["fixing_part_id"])
        for part in rec["parts"]:
            if part["part_no"]:
                refs.add(part["part_no"])
    return refs


def _precheck_p1(records: List[dict[str, Any]], ebom_conflict_parts: set[str]) -> List[Issue]:
    issues: List[Issue] = []
    refs = _referenced_parts(records)

    for part_no in sorted(refs):
        if part_no in ebom_conflict_parts:
            issues.append(
                _issue(
                    level="error",
                    message=(
                        f"零件号 {part_no} 在EBOM中存在重复定义，且 "
                        "Part Rev / Minor Rev / Applied Material / Gauge 不一致。"
                        "该零件不会参与 rule8-11 的比对。"
                    ),
                    sheet="EBOM",
                    row=None,
                    column="Part No.",
                    rule="P1",
                    group_key=f"P1:conflict:{part_no}",
                    group_title=f"零件 {part_no} 在 EBOM 中存在冲突定义",
                    entity_type="part",
                    entity_value=part_no,
                )
            )
    return issues


def _run_rule1(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []

    for rec in records:
        sheet = rec["sheet"]
        row = rec["row"]
        connect_id = rec["connect_id"]
        fixing_part_id = rec["fixing_part_id"]

        ok = True
        if len(connect_id) != 14:
            ok = False
        elif connect_id[:3] != sheet:
            ok = False
        elif not connect_id[3:9].isdigit():
            ok = False
        elif connect_id[9:10] != "_":
            ok = False
        elif not connect_id[10:14].isdigit():
            ok = False

        if not ok:
            issues.append(
                _issue(
                    level="error",
                    message=f"Connect ID 格式不符合规范：{connect_id}",
                    sheet=sheet,
                    row=row,
                    column="Connect ID",
                    rule=_display_rule_name("rule1"),
                    group_key=f"rule1:connect:{connect_id or f'{sheet}:{row}'}",
                    group_title=f"Connect ID 格式不符合规范：{connect_id or '(空值)'}",
                    entity_type="connect_id",
                    entity_value=connect_id or "",
                )
            )

        if connect_id and len(connect_id) >= 9 and fixing_part_id:
            if len(fixing_part_id) < 8 or fixing_part_id[2:8] != connect_id[3:9]:
                issues.append(
                    _issue(
                        level="error",
                        message=f"Fixing part ID 与 Connect ID 编码段不一致：{fixing_part_id} vs {connect_id}",
                        sheet=sheet,
                        row=row,
                        column="Fixing part ID",
                        rule=_display_rule_name("rule1"),
                        group_key=f"rule1:fixing_map:{fixing_part_id or f'{sheet}:{row}'}",
                        group_title=f"Fixing part ID 与 Connect ID 编码段不一致：{fixing_part_id or '(空值)'}",
                        entity_type="fixing_part",
                        entity_value=fixing_part_id or "",
                        details={
                            "fixing_part_id": fixing_part_id,
                            "connect_id": connect_id,
                            "fixing_part_id_3_8": fixing_part_id[2:8] if len(fixing_part_id) >= 8 else "",
                            "connect_id_4_9": connect_id[3:9] if len(connect_id) >= 9 else "",
                        },
                    )
                )

    return issues


def _run_rule2(records: List[dict[str, Any]], material_depot: set[str], material_depot_ready: bool) -> List[Issue]:
    issues: List[Issue] = []
    if not material_depot_ready:
        return issues

    for rec in records:
        for part in rec["parts"]:
            material = part["material"]
            if material and material not in material_depot:
                issues.append(
                    _issue(
                        level="error",
                        message=f'Part{part["slot"]} 材料牌号 {material} 不在材料库中',
                        sheet=rec["sheet"],
                        row=rec["row"],
                        column=f'PART {part["slot"]} Material',
                        rule=_display_rule_name("rule2"),
                        group_key=f"rule2:material:{material}",
                        group_title=f"材料牌号 {material} 不在材料库中",
                        entity_type="material",
                        entity_value=material,
                    )
                )
    return issues


def _run_rule3(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for rec in records:
        if rec["connect_id"]:
            grouped[rec["connect_id"]].append(rec)

    for connect_id, items in grouped.items():
        if len(items) > 1:
            for rec in items:
                issues.append(
                    _issue(
                        level="error",
                        message=f"连接点号 {connect_id} 重复",
                        sheet=rec["sheet"],
                        row=rec["row"],
                        column="Connect ID",
                        rule=_display_rule_name("rule3"),
                        group_key=f"rule3:connect_dup:{connect_id}",
                        group_title=f"连接点号 {connect_id} 重复",
                        entity_type="connect_id",
                        entity_value=connect_id,
                    )
                )
    return issues


def _run_rule4(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []

    by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for rec in records:
        by_sheet[rec["sheet"]].append(rec)

    seen_pairs = set()
    for sheet, items in by_sheet.items():
        n = len(items)
        for i in range(n):
            a = items[i]
            if a["x"] is None or a["y"] is None or a["z"] is None:
                continue
            for j in range(i + 1, n):
                b = items[j]
                if b["x"] is None or b["y"] is None or b["z"] is None:
                    continue

                dx = a["x"] - b["x"]
                dy = a["y"] - b["y"]
                dz = a["z"] - b["z"]
                euclidean = math.sqrt(dx * dx + dy * dy + dz * dz)
                manhattan = abs(dx) + abs(dy) + abs(dz)

                if euclidean <= 10 and manhattan < 100:
                    pair_ids = sorted([a["connect_id"], b["connect_id"]])
                    pair_key = tuple(sorted([f'{a["sheet"]}:{a["row"]}', f'{b["sheet"]}:{b["row"]}']))
                    if pair_key in seen_pairs:
                        continue
                    seen_pairs.add(pair_key)

                    issues.append(
                        _issue(
                            level="warning",
                            message=f'连接点位置重复，{a["connect_id"]} 与 {b["connect_id"]} 距离过近',
                            sheet=sheet,
                            row=a["row"],
                            column="X/Y/Z",
                            rule=_display_rule_name("rule4"),
                            group_key=f"rule4:pos_dup:{sheet}:{pair_ids[0]}:{pair_ids[1]}",
                            group_title=f"连接点位置重复：{pair_ids[0]} 与 {pair_ids[1]} 距离过近",
                            entity_type="connect_pair",
                            entity_value=f"{pair_ids[0]} / {pair_ids[1]}",
                            details={
                                "connect_id_a": a["connect_id"],
                                "connect_id_b": b["connect_id"],
                                "distance": f"{euclidean:.3f}",
                                "manhattan": f"{manhattan:.3f}",
                            },
                        )
                    )

    return issues


def _run_rule5(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []
    part_revs: dict[str, set[str]] = defaultdict(set)
    ref_rows: dict[str, set[tuple[str, int]]] = defaultdict(set)

    for rec in records:
        for part in rec["parts"]:
            if part["part_no"] and part["rev"]:
                part_revs[part["part_no"]].add(part["rev"])
                ref_rows[part["part_no"]].add((rec["sheet"], rec["row"]))

    for part_no, revs in part_revs.items():
        if len(revs) > 1:
            rev_text = _format_values(revs)
            for sheet, row in sorted(ref_rows[part_no]):
                issues.append(
                    _issue(
                        level="error",
                        message=f"零件 {part_no} 存在多个版本号：{rev_text}",
                        sheet=sheet,
                        row=row,
                        column="PART Rev",
                        rule=_display_rule_name("rule5"),
                        group_key=f"rule5:part_rev:{part_no}",
                        group_title=f"零件 {part_no} 存在多个版本号：{rev_text}",
                        entity_type="part",
                        entity_value=part_no,
                        details={"versions": sorted(revs)},
                    )
                )
    return issues


def _run_rule6(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []
    part_materials: dict[str, set[str]] = defaultdict(set)
    ref_rows: dict[str, set[tuple[str, int]]] = defaultdict(set)

    for rec in records:
        for part in rec["parts"]:
            if part["part_no"] and part["material"]:
                part_materials[part["part_no"]].add(part["material"])
                ref_rows[part["part_no"]].add((rec["sheet"], rec["row"]))

    for part_no, materials in part_materials.items():
        if len(materials) > 1:
            material_text = _format_values(materials)
            for sheet, row in sorted(ref_rows[part_no]):
                issues.append(
                    _issue(
                        level="error",
                        message=f"零件 {part_no} 存在多个材料牌号：{material_text}",
                        sheet=sheet,
                        row=row,
                        column="PART Material",
                        rule=_display_rule_name("rule6"),
                        group_key=f"rule6:part_material:{part_no}",
                        group_title=f"零件 {part_no} 存在多个材料牌号：{material_text}",
                        entity_type="part",
                        entity_value=part_no,
                        details={"materials": sorted(materials)},
                    )
                )
    return issues


def _run_rule7(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []
    part_gauges: dict[str, set[str]] = defaultdict(set)
    ref_rows: dict[str, set[tuple[str, int]]] = defaultdict(set)
    part_has_dc: dict[str, bool] = defaultdict(bool)

    for rec in records:
        for part in rec["parts"]:
            if part["gauge_num"] is not None and part["gauge_num"] < 0:
                issues.append(
                    _issue(
                        level="error",
                        message=f'Part{part["slot"]} 零件厚度为负值：{part["gauge_raw"]}',
                        sheet=rec["sheet"],
                        row=rec["row"],
                        column=f'PART {part["slot"]} Gauge',
                        rule=_display_rule_name("rule7"),
                        group_key=f"rule7:negative_gauge:{rec['sheet']}:{rec['row']}:{part['slot']}",
                        group_title="存在负厚度数据",
                        entity_type="gauge",
                        entity_value=part["gauge_raw"],
                        details={
                            "part_no": part["part_no"],
                            "slot": part["slot"],
                            "gauge": part["gauge_raw"],
                        },
                    )
                )

            if part["part_no"]:
                if "DC" in part["material"]:
                    part_has_dc[part["part_no"]] = True

                if part["gauge_raw"]:
                    part_gauges[part["part_no"]].add(_norm_num_text(part["gauge_raw"]))
                    ref_rows[part["part_no"]].add((rec["sheet"], rec["row"]))

    for part_no, gauges in part_gauges.items():
        if part_has_dc.get(part_no):
            continue
        if len(gauges) > 1:
            gauge_text = _format_values(gauges)
            for sheet, row in sorted(ref_rows[part_no]):
                issues.append(
                    _issue(
                        level="error",
                        message=f"零件 {part_no} 存在多个厚度：{gauge_text}",
                        sheet=sheet,
                        row=row,
                        column="PART Gauge",
                        rule=_display_rule_name("rule7"),
                        group_key=f"rule7:part_gauge:{part_no}",
                        group_title=f"零件 {part_no} 存在多个厚度：{gauge_text}",
                        entity_type="part",
                        entity_value=part_no,
                        details={"gauges": sorted(gauges)},
                    )
                )
    return issues


def _run_rule8_to_11(
    records: List[dict[str, Any]],
    ebom_uploaded: bool,
    ebom_ready: bool,
    bom_lookup: dict[str, dict[str, Any]],
    conflict_parts: set[str],
) -> List[Issue]:
    issues: List[Issue] = []

    if not ebom_uploaded:
        # 未上传 EBOM 的提示已在 _data_maturity_check 中统一输出，避免重复信息。
        return issues

    if not ebom_ready:
        issues.append(
            _issue(
                level="warning",
                message="EBOM 结构异常或读取失败，rule8-11 已跳过。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("rule8-11"),
                group_key="rule8-11:ebom_invalid",
                group_title="EBOM 结构异常，rule8-11 已跳过",
                entity_type="ebom",
            )
        )
        return issues

    for rec in records:
        fixing_part_id = rec["fixing_part_id"]
        fixing_part_rev = rec["fixing_part_rev"]

        # rule8
        if fixing_part_id:
            if fixing_part_id in conflict_parts:
                pass
            else:
                bom = bom_lookup.get(fixing_part_id)
                if not bom:
                    issues.append(
                        _issue(
                            level="error",
                            message=f"Fixing Part {fixing_part_id} 在BOM中不存在",
                            sheet=rec["sheet"],
                            row=rec["row"],
                            column="Fixing part ID",
                            rule=_display_rule_name("rule8"),
                            group_key=f"rule8:no_bom:{fixing_part_id}",
                            group_title=f"Fixing Part {fixing_part_id} 在BOM中不存在",
                            entity_type="fixing_part",
                            entity_value=fixing_part_id,
                        )
                    )
                else:
                    if fixing_part_rev != bom["version"]:
                        issues.append(
                            _issue(
                                level="error",
                                message=(
                                    f"Fixing Part {fixing_part_id} 版本号与BOM不一致"
                                    f"（点表={fixing_part_rev}，BOM={bom['version']}）"
                                ),
                                sheet=rec["sheet"],
                                row=rec["row"],
                                column="Fixing part Rev",
                                rule=_display_rule_name("rule8"),
                                group_key=f"rule8:version_mismatch:{fixing_part_id}",
                                group_title=f"Fixing Part {fixing_part_id} 版本号与BOM不一致",
                                entity_type="fixing_part",
                                entity_value=fixing_part_id,
                                details={
                                    "point_version": fixing_part_rev,
                                    "bom_version": bom["version"],
                                },
                            )
                        )

        for part in rec["parts"]:
            part_no = part["part_no"]
            if not part_no:
                continue

            if part_no in conflict_parts:
                continue

            bom = bom_lookup.get(part_no)
            if not bom:
                issues.append(
                    _issue(
                        level="error",
                        message=f'Part{part["slot"]} {part_no} 在BOM中不存在',
                        sheet=rec["sheet"],
                        row=rec["row"],
                        column=f'PART {part["slot"]}',
                        rule=_display_rule_name("rule9"),
                        group_key=f"rule9:no_bom:{part_no}",
                        group_title=f"零件 {part_no} 在BOM中不存在",
                        entity_type="part",
                        entity_value=part_no,
                        details={"slot": part["slot"]},
                    )
                )
                continue

            # rule9
            if part["rev"] != bom["version"]:
                issues.append(
                    _issue(
                        level="error",
                        message=(
                            f'Part{part["slot"]} {part_no} 版本号与BOM不一致'
                            f'（点表={part["rev"]}，BOM={bom["version"]}）'
                        ),
                        sheet=rec["sheet"],
                        row=rec["row"],
                        column=f'PART {part["slot"]} Rev',
                        rule=_display_rule_name("rule9"),
                        group_key=f"rule9:version_mismatch:{part_no}",
                        group_title=f"零件 {part_no} 版本号与BOM不一致",
                        entity_type="part",
                        entity_value=part_no,
                        details={
                            "slot": part["slot"],
                            "point_version": part["rev"],
                            "bom_version": bom["version"],
                        },
                    )
                )

            # rule10
            if part["material"] != bom["applied_material"]:
                issues.append(
                    _issue(
                        level="error",
                        message=(
                            f'Part{part["slot"]} {part_no} 材料与BOM不一致'
                            f'（点表={part["material"]}，BOM={bom["applied_material"]}）'
                        ),
                        sheet=rec["sheet"],
                        row=rec["row"],
                        column=f'PART {part["slot"]} Material',
                        rule=_display_rule_name("rule10"),
                        group_key=f"rule10:material_mismatch:{part_no}",
                        group_title=f"零件 {part_no} 材料与BOM不一致",
                        entity_type="part",
                        entity_value=part_no,
                        details={
                            "slot": part["slot"],
                            "point_material": part["material"],
                            "bom_material": bom["applied_material"],
                        },
                    )
                )

            # rule11
            if "DC" not in part["material"]:
                gauge_text = _norm_num_text(part["gauge_raw"])
                if gauge_text and gauge_text not in bom["gauge_options"]:
                    issues.append(
                        _issue(
                            level="error",
                            message=(
                                f'Part{part["slot"]} {part_no} 厚度与BOM不一致'
                                f'（点表={gauge_text}，BOM={bom["gauge_raw"]}）'
                            ),
                            sheet=rec["sheet"],
                            row=rec["row"],
                            column=f'PART {part["slot"]} Gauge',
                            rule=_display_rule_name("rule11"),
                            group_key=f"rule11:gauge_mismatch:{part_no}",
                            group_title=f"零件 {part_no} 厚度与BOM不一致",
                            entity_type="part",
                            entity_value=part_no,
                            details={
                                "slot": part["slot"],
                                "point_gauge": gauge_text,
                                "bom_gauge": bom["gauge_raw"],
                                "bom_gauge_options": bom["gauge_options"],
                            },
                        )
                    )

    return issues


def _run_rule12(records: List[dict[str, Any]]) -> List[Issue]:
    issues: List[Issue] = []

    for rec in records:
        gauge_count = sum(1 for p in rec["parts"] if _normalize_text(p["gauge_raw"]))
        is_four_layer = gauge_count == 4
        is_indirect_rsw = rec["extra_info"] == "indirect RSW"
        is_process_joint = rec["process_joint"] == "Process Joint"

        if is_four_layer and not is_process_joint:
            issues.append(
                _issue(
                    level="error",
                    message="四层板连接点必须标记为 Process Joint",
                    sheet=rec["sheet"],
                    row=rec["row"],
                    column="Process Joint",
                    rule=_display_rule_name("rule12"),
                    group_key="rule12:four_layer_missing_process_joint",
                    group_title="四层板连接点未标记为 Process Joint",
                    entity_type="process_joint",
                )
            )

        if is_indirect_rsw and not is_process_joint:
            issues.append(
                _issue(
                    level="error",
                    message="单边点焊（indirect RSW）必须标记为 Process Joint",
                    sheet=rec["sheet"],
                    row=rec["row"],
                    column="Process Joint",
                    rule=_display_rule_name("rule12"),
                    group_key="rule12:indirect_rsw_missing_process_joint",
                    group_title="indirect RSW 未标记为 Process Joint",
                    entity_type="process_joint",
                )
            )

        if is_process_joint and not (is_four_layer or is_indirect_rsw):
            issues.append(
                _issue(
                    level="error",
                    message="该连接点不属于单边点焊或四层搭接，请确认是否为工艺点",
                    sheet=rec["sheet"],
                    row=rec["row"],
                    column="Process Joint",
                    rule=_display_rule_name("rule12"),
                    group_key="rule12:process_joint_without_reason",
                    group_title="存在无合理依据的 Process Joint",
                    entity_type="process_joint",
                )
            )

    return issues


def _data_maturity_check(
    file_bytes: bytes,
    filename: str,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    issues: List[Issue] = []

    try:
        main_wb = _load_workbook_from_bytes(file_bytes)
    except Exception as exc:
        return [
            _issue(
                level="error",
                message=f"主文件读取失败：{exc}",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("input"),
                group_key="input:main_read_failed",
                group_title="主文件读取失败",
                entity_type="main_file",
                details={"error": str(exc)},
            )
        ]

    all_records, read_issues, existing_sheets = _collect_all_records(main_wb)
    issues.extend(read_issues)

    if not existing_sheets:
        issues.append(
            _issue(
                level="error",
                message="主文件中未找到 RSW / SPR / FDS 三个目标 Sheet。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("input"),
                group_key="input:target_sheet_missing",
                group_title="主文件缺少目标 Sheet（RSW / SPR / FDS）",
                entity_type="sheet",
            )
        )
        return issues

    material_depot, material_issues = _load_material_depot()
    issues.extend(material_issues)
    material_depot_ready = len(material_depot) > 0

    ebom_uploaded = ebom_bytes is not None
    ebom_ready = False
    bom_lookup: dict[str, dict[str, Any]] = {}
    conflict_parts: set[str] = set()

    if ebom_uploaded:
        bom_lookup, ebom_conflict_parts, ebom_issues = _load_ebom(ebom_bytes)
        issues.extend(ebom_issues)
        ebom_ready = not any(x.level == "error" for x in ebom_issues)
        if ebom_ready:
            conflict_parts = ebom_conflict_parts
            issues.extend(_precheck_p1(all_records, conflict_parts))
    else:
        issues.append(
            _issue(
                level="info",
                message="未上传对应产品 EBOM 清单，仅执行不依赖 BOM 的规则。",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("input"),
                group_key="input:no_ebom_uploaded",
                group_title="未上传 EBOM，仅执行非 BOM 相关规则",
                entity_type="ebom",
            )
        )

    records_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in all_records:
        records_by_sheet[record["sheet"]].append(record)

    for sheet_name in existing_sheets:
        sheet_records = records_by_sheet.get(sheet_name, [])
        issues.extend(_run_rule1(sheet_records))
        issues.extend(_run_rule3(sheet_records))
        issues.extend(_run_rule4(sheet_records))
        issues.extend(_run_rule12(sheet_records))

    issues.extend(_run_rule2(all_records, material_depot, material_depot_ready))
    issues.extend(_run_rule5(all_records))
    issues.extend(_run_rule6(all_records))
    issues.extend(_run_rule7(all_records))

    issues.extend(_run_rule8_to_11(all_records, ebom_uploaded, ebom_ready, bom_lookup, conflict_parts))

    if ebom_uploaded and ebom_filename and ebom_ready:
        issues.append(
            _issue(
                level="info",
                message=f"本次数据成熟度检查已使用 EBOM 文件：{ebom_filename}",
                sheet=None,
                row=None,
                column=None,
                rule=_display_rule_name("input"),
                group_key="input:ebom_used",
                group_title="本次检查已使用 EBOM 文件",
                entity_type="ebom_file",
                entity_value=ebom_filename,
            )
        )

    return issues


def _required_fields_demo(
    _: bytes,
    filename: str,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    return [
        Issue(
            level="error",
            message="客户名称不能为空",
            sheet="客户清单",
            row=3,
            column="B",
            rule="required_fields_demo",
            group_key="required_fields_demo:customer_name_empty",
            group_title="客户名称为空",
        ),
        Issue(
            level="error",
            message="订单编号不能为空",
            sheet="订单明细",
            row=8,
            column="A",
            rule="required_fields_demo",
            group_key="required_fields_demo:order_no_empty",
            group_title="订单编号为空",
        ),
    ]


def _date_format_demo(
    _: bytes,
    filename: str,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    return [
        Issue(
            level="warning",
            message="日期格式建议统一为 YYYY-MM-DD",
            sheet="订单明细",
            row=5,
            column="D",
            rule="date_format_demo",
            group_key="date_format_demo:format",
            group_title="日期格式建议统一",
        ),
        Issue(
            level="warning",
            message="检测到文本日期，建议转换为标准日期单元格",
            sheet="发票记录",
            row=12,
            column="C",
            rule="date_format_demo",
            group_key="date_format_demo:text_date",
            group_title="存在文本日期",
        ),
    ]


def _enum_value_demo(
    _: bytes,
    filename: str,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    return [
        Issue(
            level="info",
            message="状态值“已完成-手工”不在推荐枚举内，建议映射为“已完成”",
            sheet="任务列表",
            row=6,
            column="F",
            rule="enum_value_demo",
            group_key="enum_value_demo:status",
            group_title="存在非常规状态值",
        )
    ]


def _cross_sheet_demo(
    _: bytes,
    filename: str,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    return [
        Issue(
            level="error",
            message="子表中的客户ID在主表中不存在",
            sheet="订单明细",
            row=9,
            column="B",
            rule="cross_sheet_demo",
            group_key="cross_sheet_demo:customer_id",
            group_title="子表客户ID不存在",
        )
    ]


def _range_check_demo(
    _: bytes,
    filename: str,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    return [
        Issue(
            level="warning",
            message="折扣率超过 100%，请确认数据是否正确",
            sheet="促销策略",
            row=7,
            column="E",
            rule="range_check_demo",
            group_key="range_check_demo:discount",
            group_title="折扣率超过 100%",
        )
    ]


RuleHandler = Callable[[bytes, str, bytes | None, str | None], List[Issue]]

RULE_HANDLERS: Dict[str, RuleHandler] = {
    "data_maturity_check": _data_maturity_check,
    "required_fields_demo": _required_fields_demo,
    "date_format_demo": _date_format_demo,
    "enum_value_demo": _enum_value_demo,
    "cross_sheet_demo": _cross_sheet_demo,
    "range_check_demo": _range_check_demo,
}


def get_default_enabled_rule_ids() -> List[str]:
    return [r["id"] for r in get_rules() if r.get("enabled")]


def sanitize_selected_rule_ids(selected_rule_ids: List[str]) -> List[str]:
    valid = set(RULE_HANDLERS.keys())
    return [rule_id for rule_id in selected_rule_ids if rule_id in valid]


def run_checks(
    file_bytes: bytes,
    filename: str,
    selected_rule_ids: List[str] | None = None,
    ebom_bytes: bytes | None = None,
    ebom_filename: str | None = None,
) -> List[Issue]:
    if not selected_rule_ids:
        selected_rule_ids = get_default_enabled_rule_ids()

    selected_rule_ids = sanitize_selected_rule_ids(selected_rule_ids)

    issues: List[Issue] = []

    for rule_id in selected_rule_ids:
        handler = RULE_HANDLERS.get(rule_id)
        if not handler:
            continue

        try:
            issues.extend(handler(file_bytes, filename, ebom_bytes, ebom_filename))
        except Exception:
            logger.exception("Rule execution failed: %s", rule_id)
            issues.append(
                Issue(
                    level="error",
                    message=f"规则执行异常：{rule_id}",
                    sheet=None,
                    row=None,
                    column=None,
                    rule=rule_id,
                    group_key=f"{rule_id}:execution_failed",
                    group_title=f"规则执行异常：{rule_id}",
                )
            )

    return issues
